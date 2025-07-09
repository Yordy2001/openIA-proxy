import pandas as pd
import openpyxl
from typing import List, Dict, Any, Union
import io
import os
from fastapi import HTTPException
from models import ExcelStructuredData, ExcelSheetData
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side


class ExcelProcessor:
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls'}
    
    def validate_file(self, filename: str, file_size: int, max_size: int) -> bool:
        """Validate if the file is supported and within size limits"""
        file_ext = os.path.splitext(filename)[1].lower()
        
        if file_ext not in self.supported_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Formato de archivo no soportado. Formatos permitidos: {', '.join(self.supported_extensions)}"
            )
        
        if file_size > max_size:
            raise HTTPException(
                status_code=400,
                detail=f"Archivo demasiado grande. Tamaño máximo: {max_size / (1024*1024):.1f}MB"
            )
        
        return True
    
    def extract_structured_data(self, file_content: bytes, filename: str) -> ExcelStructuredData:
        """Extract structured data from Excel file for table display"""
        try:
            excel_file = pd.ExcelFile(io.BytesIO(file_content))
            sheets = []
            
            for sheet_name in excel_file.sheet_names:
                try:
                    # Read the sheet with headers
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    # Clean column names
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    # Remove completely empty rows
                    df = df.dropna(how='all')
                    
                    if df.empty:
                        continue
                    
                    # Fill NaN values with empty strings for display
                    df_display = df.fillna("")
                    
                    # Convert datetime columns to strings
                    for col in df_display.columns:
                        if df_display[col].dtype == 'datetime64[ns]':
                            df_display[col] = df_display[col].astype(str)
                    
                    # Identify numeric columns
                    numeric_columns = list(df.select_dtypes(include=['number']).columns)
                    
                    # Create sheet data
                    sheet_data = ExcelSheetData(
                        sheet_name=sheet_name,
                        columns=list(df_display.columns),
                        rows=df_display.to_dict('records'),
                        shape=list(df_display.shape),
                        numeric_columns=numeric_columns
                    )
                    
                    sheets.append(sheet_data)
                    
                except Exception as e:
                    print(f"Error processing sheet '{sheet_name}': {e}")
                    continue
            
            return ExcelStructuredData(
                filename=filename,
                sheets=sheets,
                metadata={
                    "total_sheets": len(sheets),
                    "processed_at": pd.Timestamp.now().isoformat()
                }
            )
            
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error al procesar el archivo Excel: {str(e)}"
            )
    
    def update_cell_data(self, structured_data: ExcelStructuredData, sheet_name: str, 
                        row: int, column: str, value: Union[str, int, float, None]) -> ExcelStructuredData:
        """Update a specific cell in the structured data"""
        try:
            # Find the sheet
            sheet_index = None
            for i, sheet in enumerate(structured_data.sheets):
                if sheet.sheet_name == sheet_name:
                    sheet_index = i
                    break
            
            if sheet_index is None:
                raise HTTPException(
                    status_code=404,
                    detail=f"Hoja '{sheet_name}' no encontrada"
                )
            
            # Update the cell
            if row < len(structured_data.sheets[sheet_index].rows):
                if column in structured_data.sheets[sheet_index].columns:
                    # Convert value to appropriate type
                    if isinstance(value, str) and value.strip() == "":
                        value = None
                    elif isinstance(value, str):
                        # Try to convert to number if it's a numeric column
                        if column in structured_data.sheets[sheet_index].numeric_columns:
                            try:
                                value = float(value) if '.' in value else int(value)
                            except ValueError:
                                pass  # Keep as string if conversion fails
                    
                    structured_data.sheets[sheet_index].rows[row][column] = value
                else:
                    raise HTTPException(
                        status_code=404,
                        detail=f"Columna '{column}' no encontrada"
                    )
            else:
                raise HTTPException(
                    status_code=404,
                    detail=f"Fila {row} no encontrada"
                )
            
            return structured_data
            
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error al actualizar celda: {str(e)}"
            )
    
    def export_to_excel(self, structured_data: ExcelStructuredData) -> bytes:
        """Export structured data back to Excel bytes"""
        try:
            output = io.BytesIO()
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            if wb.active is not None:
                wb.remove(wb.active)  # Remove the default sheet
            
            for sheet in structured_data.sheets:
                # Convert rows back to DataFrame
                df = pd.DataFrame(sheet.rows, columns=sheet.columns)
                
                # Create worksheet
                ws = wb.create_sheet(title=sheet.sheet_name)
                
                # Write headers
                for col_idx, column in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Write data
                for row_idx, row in enumerate(df.to_dict('records'), 2):
                    for col_idx, column in enumerate(df.columns, 1):
                        value = row[column]
                        # Handle None values
                        if pd.isna(value):
                            value = ""
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border
            
            # Save to BytesIO
            wb.save(output)
            output.seek(0)
            return output.read()
            
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error al exportar Excel: {str(e)}"
            )
    
    def extract_data_from_excel(self, file_content: bytes, filename: str) -> str:
        """Extract and format data from Excel file for AI analysis"""
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(io.BytesIO(file_content))
            
            formatted_data = []
            formatted_data.append(f"=== ANÁLISIS DE ARCHIVO: {filename} ===\n")
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                formatted_data.append(f"\n--- HOJA: {sheet_name} ---")
                
                try:
                    # Read the sheet
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    # Remove completely empty rows and columns
                    df = df.dropna(how='all').dropna(axis=1, how='all')
                    
                    if df.empty:
                        formatted_data.append("Esta hoja está vacía o no contiene datos válidos.")
                        continue
                    
                    # Add sheet information
                    formatted_data.append(f"Dimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")
                    
                    # Convert to string representation
                    formatted_data.append("\nDatos:")
                    
                    # Process each row with row numbers
                    for idx, row in df.iterrows():
                        row_data = []
                        for col_idx, value in enumerate(row):
                            if pd.notna(value):
                                # Format numbers appropriately
                                if isinstance(value, (int, float)):
                                    if isinstance(value, float) and value.is_integer():
                                        formatted_value = f"{int(value)}"
                                    else:
                                        formatted_value = f"{value:,.2f}"
                                else:
                                    formatted_value = str(value)
                                row_data.append(f"Col{col_idx+1}: {formatted_value}")
                        
                        if row_data:
                            formatted_data.append(f"Fila {idx+1}: {' | '.join(row_data)}")
                    
                    # Add summary statistics for numeric columns
                    numeric_columns = df.select_dtypes(include=['number']).columns
                    if len(numeric_columns) > 0:
                        formatted_data.append(f"\nResumen estadístico para columnas numéricas:")
                        for col in numeric_columns:
                            col_data = df[col].dropna()
                            if len(col_data) > 0:
                                formatted_data.append(
                                    f"  Columna {col+1}: Suma={col_data.sum():,.2f}, "
                                    f"Promedio={col_data.mean():,.2f}, "
                                    f"Min={col_data.min():,.2f}, "
                                    f"Max={col_data.max():,.2f}"
                                )
                    
                except Exception as e:
                    formatted_data.append(f"Error al procesar la hoja '{sheet_name}': {str(e)}")
                    continue
            
            return "\n".join(formatted_data)
            
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error al procesar el archivo Excel: {str(e)}"
            )
    
    def process_multiple_files(self, files_data: List[Dict[str, Any]]) -> str:
        """Process multiple Excel files and combine their data"""
        all_data = []
        
        for file_data in files_data:
            filename = file_data['filename']
            content = file_data['content']
            
            try:
                file_analysis = self.extract_data_from_excel(content, filename)
                all_data.append(file_analysis)
                all_data.append("\n" + "="*80 + "\n")
            except Exception as e:
                all_data.append(f"Error procesando {filename}: {str(e)}")
                all_data.append("\n" + "="*80 + "\n")
        
        return "\n".join(all_data) 